
import openpyxl

wb = openpyxl.load_workbook("octobre.xlsx")
print(wb.sheetnames)

# sheet = wb["Feuil1"]
# sheet = wb[wb.sheetnames[0]]
sheet = wb.active

print(sheet.max_row)
print(sheet.max_column)

# cell = sheet["A1"]
for row in range(2,7):
    cell = sheet.cell(row,2) #(nb colonne, nb ligne)
    print(cell.value)